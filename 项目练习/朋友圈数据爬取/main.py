import time
import uiautomation as auto
from openpyxl import Workbook,load_workbook
import tkinter as tk

def append_to_excel(filename,sheetname,data):
	try:
		#打开现有的excel
		workbook = load_workbook(filename)
		#选择要早错的工作表
		sheet = workbook[sheetname]
	except:
		# 如果文件不存在，则创建一个新的工作簿
		workbook = Workbook()
		sheet = workbook.active
		sheet.title = sheetname
	
	#追加数据到工作表
	name = data[0]
	sheet.append([name])
	#保存
	workbook.save(filename)

""" 
def button_clikc():
	user_input = entry.get()


#创建窗口
window = tk.TK()
window.title("微信朋友圈")		
window.geometry("500x700")

entry = tk.Entry(window)
entry.pack()
 """

wechatWindow = auto.WindowControl(searchDepth=1, Name="微信", ClassName='WeChatMainWndForPC')

#打开朋友圈
button = wechatWindow.ButtonControl(Name ="朋友圈")
button.Click()

#构造朋友圈窗口
friendWindow = auto.WindowControl(searchDepth = 1, Name="朋友圈", ClassName="SnsWnd")

#获取朋友圈窗口聚点
friendWindow.SetFocus()
#获取朋友圈的列表
listControls = friendWindow.ListControl(Name='朋友圈')

#Excel基础数据
filename = 'test.xlsx'
sheetname = 'Sheet1'

flag = True
processed_posts = set()

while flag:
	
	for item in listControls.GetChildren():
        # 获取到用户名和文案
		content = item.Name.strip()
		content_split = content.split("\n")

		# 把content_split列表里的信息提取出来
		nick_name = content_split[0]
		date = content_split[-1]

		# 设置时间让pyq内容加载(不设置的话可能会漏掉一些朋友圈)
		time.sleep(5)
		# 爬取两天前的pyq
		if date != "1小时前":
			item.SendKeys("{Down}")
		else:
			flag = False

		# Check if the post contains the text "测试"
		if "​测试" in content:
			# Check if the post has already been processed
			if content not in processed_posts:
				# 保存到excel
				append_to_excel(filename, sheetname, content_split)
				processed_posts.add(content)

		item = listControls.GetLastChildControl()
		print(content_split)
		 

	


		


